from tkinter import ttk
import tkinter as tk
import customtkinter as ck

from openpyxl import load_workbook, Workbook, styles
import os 

from qrcodem.generate import employee_qr


base_path = os.path.dirname(__file__)


try:
    with open(os.path.join(base_path, 'pay.xlsx'), 'rb') as f:
        wb = load_workbook(f)
except FileNotFoundError:
    wb = Workbook()
    wb.save(os.path.join(base_path, 'pay.xlsx'))
    wb = load_workbook(os.path.join(base_path, 'pay.xlsx'))

class AttendaceFrame(ck.CTkFrame):
    def __init__(self, master):
        super().__init__(master, width=100, height=400)
        self.pack(fill=tk.BOTH, expand=True)
        # self.pack()
        self.create_widgets()

        self.time = tk.StringVar()

        self.time_out_val = tk.StringVar()

        



    def create_widgets(self):
        self.employee = ttk.LabelFrame(self, text="Employee")
        self.employee.pack(fill=tk.BOTH)
        self.employee.columnconfigure(0, weight=1)

        self.employee_id = ck.CTkEntry(self)
        self.employee_id.pack()

        self.employee_id_label = ck.CTkLabel(self, text="Employee ID")
        self.employee_id_label.pack()

        self.employee_name = ck.CTkEntry(self)
        self.employee_name.pack()
        self.employee_name_label = ck.CTkLabel(self, text="Employee Name")
        self.employee_name_label.pack()


        self.time_in = ck.CTkEntry(self)
        self.time_in.pack()
        self.time_in_label = ck.CTkLabel(self, text="Time In")
        self.time_in_label.pack()

        self.log_out = ck.CTkEntry(self)
        self.log_out.pack()

        self.log_out_label = ck.CTkLabel(self, text="Log Out")
        self.log_out_label.pack()
    wb.save(os.path.join(base_path, 'pay.xlsx'))


        

class EmpployeeFrame(ck.CTkFrame):
    def __init__(self, master):
        super().__init__(master, width=400, height=400)
        self.pack(fill=tk.BOTH, expand=True)
        # self.pack()

    
class GenerateId(ck.CTkFrame):
    count = 0
    def __init__(self, master):
        GenerateId.count += 1
        super().__init__(master, width=400, height=400)
        self.pack(fill=tk.BOTH,)
        # self.pack()
        self.firstname = tk.StringVar()
        self.lastname = tk.StringVar()
        self.id_number = tk.StringVar()     
        self.create_widgets()


    
    def create_widgets(self):
        self.firstname_label = ck.CTkLabel(self, text="First Name")
        self.firstname_label.pack()

        self.firstname_entry = ck.CTkEntry(self, textvariable=self.firstname)
        self.firstname_entry.pack()

        self.lastname_label = ck.CTkLabel(self, text="Last Name")
        self.lastname_label.pack()

        self.lastname_entry = ck.CTkEntry(self, textvariable=self.lastname)
        self.lastname_entry.pack()

        self.create = ck.CTkButton(self, text="Create ID", command=self.create_id)
        self.create.pack()


    def create_id(self):
            print(self.firstname.get() + self.lastname.get())
            employee_qr(name=self.firstname.get() + self.lastname.get(),  id_number=GenerateId.count)